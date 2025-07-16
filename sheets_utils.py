import os
import json
import gspread
from googleapiclient.discovery import build
from datetime import datetime, timedelta
import tempfile
from googleapiclient.http import MediaFileUpload
from dotenv import load_dotenv
load_dotenv()
import pandas as pd
from openpyxl import load_workbook
import pickle

SCOPES = [
    'https://www.googleapis.com/auth/spreadsheets',
    'https://www.googleapis.com/auth/drive'
]

# OAuth2 user authentication
with open('token.pickle', 'rb') as token:
    creds = pickle.load(token)

gc = gspread.authorize(creds)
drive_service = build('drive', 'v3', credentials=creds)

SPREADSHEET_ID = os.getenv('SPREADSHEET_ID')
SPREADSHEET_NAME = os.getenv('SPREADSHEET_NAME', 'TeamUpdates')
BACKUP_FOLDER_NAME = os.getenv('BACKUP_FOLDER_NAME', 'Backups')
BACKUP_FOLDER_ID = os.getenv('GOOGLE_BACKUP_FOLDER_ID')
OTHER_FOLDER_ID = os.getenv('GOOGLE_OTHER_FOLDER_ID')


def get_or_create_spreadsheet():
    global SPREADSHEET_ID
    if SPREADSHEET_ID:
        try:
            gc.open_by_key(SPREADSHEET_ID)
            return SPREADSHEET_ID
        except Exception:
            pass
    sh = gc.create(SPREADSHEET_NAME)
    SPREADSHEET_ID = sh.id
    return SPREADSHEET_ID


def get_or_create_user_sheet(spreadsheet_id, username):
    sh = gc.open_by_key(spreadsheet_id)
    try:
        worksheet = sh.worksheet(username)
    except gspread.exceptions.WorksheetNotFound:
        worksheet = sh.add_worksheet(title=username, rows="1000", cols="3")
        worksheet.append_row(["Date", "Time", "Update Text"])
    return worksheet


def append_update(username, update_text):
    spreadsheet_id = get_or_create_spreadsheet()
    worksheet = get_or_create_user_sheet(spreadsheet_id, username)
    now = datetime.now()
    worksheet.append_row([
        now.strftime('%Y-%m-%d'),
        now.strftime('%H:%M:%S'),
        update_text
    ])


def get_drive_file_id_by_name(name, mime_type=None, parent=None):
    query = f"name='{name}' and trashed=false"
    if mime_type:
        query += f" and mimeType='{mime_type}'"
    if parent:
        query += f" and '{parent}' in parents"
    results = drive_service.files().list(q=query, fields="files(id, name)").execute()
    files = results.get('files', [])
    return files[0]['id'] if files else None


def get_or_create_backup_folder():
    folder_id = get_drive_file_id_by_name(BACKUP_FOLDER_NAME, mime_type='application/vnd.google-apps.folder')
    if folder_id:
        return folder_id
    file_metadata = {
        'name': BACKUP_FOLDER_NAME,
        'mimeType': 'application/vnd.google-apps.folder'
    }
    folder = drive_service.files().create(body=file_metadata, fields='id').execute()
    return folder.get('id')


def export_and_backup_spreadsheet():
    spreadsheet_id = get_or_create_spreadsheet()
    sh = gc.open_by_key(spreadsheet_id)
    spreadsheet_name = sh.title
    # Find Google Sheets file in Drive
    file_id = get_drive_file_id_by_name(spreadsheet_name, mime_type='application/vnd.google-apps.spreadsheet')
    if not file_id:
        raise Exception('Spreadsheet not found in Drive')
    # Export as XLSX
    request = drive_service.files().export_media(fileId=file_id, mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    today = datetime.now().strftime('%Y-%m-%d')
    backup_filename = f"{today}_{spreadsheet_name}.xlsx"
    backup_folder_id = BACKUP_FOLDER_ID or get_or_create_backup_folder()
    with tempfile.NamedTemporaryFile(delete=False) as tmp_file:
        tmp_file.write(request.execute())
        tmp_file.flush()
        # Upload to backup folder
        file_metadata = {
            'name': backup_filename,
            'parents': [backup_folder_id]
        }
        media = MediaFileUpload(tmp_file.name, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        drive_service.files().create(body=file_metadata, media_body=media, fields='id').execute()
    os.unlink(tmp_file.name)
    delete_old_backups(backup_folder_id, spreadsheet_name)
    return backup_filename


def delete_old_backups(folder_id, spreadsheet_name):
    results = drive_service.files().list(q=f"'{folder_id}' in parents and trashed=false and name contains '{spreadsheet_name}'", fields="files(id, name, createdTime)").execute()
    files = results.get('files', [])
    now = datetime.now()
    for file in files:
        try:
            file_date = datetime.strptime(file['name'][:10], '%Y-%m-%d')
            if (now - file_date).days > 7:
                drive_service.files().delete(fileId=file['id']).execute()
        except Exception:
            continue 

LOCAL_EXCEL_FILE = 'local_updates.xlsx'

def push_local_updates_to_gsheets():
    spreadsheet_id = get_or_create_spreadsheet()
    try:
        book = load_workbook(LOCAL_EXCEL_FILE)
        for username in book.sheetnames:
            df = pd.read_excel(LOCAL_EXCEL_FILE, sheet_name=username)
            for _, row in df.iterrows():
                # Append each row to the corresponding user sheet in Google Sheets
                worksheet = get_or_create_user_sheet(spreadsheet_id, username)
                worksheet.append_row([
                    str(row['Date']),
                    str(row['Time']),
                    str(row['Update Text'])
                ])
    except FileNotFoundError:
        pass  # No local updates to push 

def upload_file_to_other_folder(local_file_path, filename=None, mime_type=None):
    """Upload a file to the folder specified by OTHER_FOLDER_ID, replacing if it already exists."""
    folder_id = OTHER_FOLDER_ID
    if not folder_id:
        raise Exception("GOOGLE_OTHER_FOLDER_ID is not set in the environment.")
    if filename is None:
        filename = os.path.basename(local_file_path)
    if mime_type is None:
        import mimetypes
        mime_type = mimetypes.guess_type(local_file_path)[0] or 'application/octet-stream'
    print(f"Uploading file {filename} to folder {folder_id} with MIME type {mime_type}")
    # Search for existing file with the same name in the folder
    query = f"'{folder_id}' in parents and name='{filename}' and trashed=false"
    results = drive_service.files().list(q=query, fields="files(id, name)").execute()
    files = results.get('files', [])
    media = MediaFileUpload(local_file_path, mimetype=mime_type, resumable=True)
    if files:
        # Replace (update) the existing file
        file_id = files[0]['id']
        drive_service.files().update(
            fileId=file_id,
            media_body=media,
            fields='id'
        ).execute()
    else:
        # Create new file in the folder
        file_metadata = {
            'name': filename,
            'parents': [folder_id]
        }
        drive_service.files().create(
            body=file_metadata,
            media_body=media,
            fields='id'
        ).execute() 